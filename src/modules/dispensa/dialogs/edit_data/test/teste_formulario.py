import pandas as pd
from openpyxl import load_workbook

def teste_carregar_formulario(file_path, df_dados):
    try:
        print("DataFrame antes de carregar o formulário:")
        print(df_dados)

        wb = load_workbook(file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
            coluna_legivel = row[0]
            valor = row[1]
            print(f"Atualizando coluna '{coluna_legivel}' com valor: {valor}")
            df_dados.at[0, coluna_legivel] = valor

        print("DataFrame após carregar o formulário:")
        print(df_dados)
    except Exception as e:
        print(f"Erro ao carregar formulário: {str(e)}")

df_dados = pd.DataFrame({'nup': ['x'], 'material_servico': ['Material']})  # Coloque as colunas iniciais
teste_carregar_formulario("C:/Users/Guilherme/Documents/formulariotestes.xlsx", df_dados)
